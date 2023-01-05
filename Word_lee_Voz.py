

import openpyxl
from openpyxl.utils import get_column_letter

# Abrir el libro de trabajo
wb = openpyxl.load_workbook('plan_contable.xlsx')

# Seleccionar la hoja de trabajo activa
ws = wb.active

# Iterar sobre cada fila y celda
for fila in ws.iter_rows():
    for celda in fila:
        # Imprimir el resultado de la función get_column_letter() junto con el valor de cada celda
        print(get_column_letter(celda.col_idx) + str(celda.row) + ": " + str(celda.value))
